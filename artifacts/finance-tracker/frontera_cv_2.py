import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import cvxpy as cp

wb = load_workbook('PORTFOLIO.xlsx', data_only=True)
ws = wb['PORT C-V']

tickers = [ws.cell(row=2, column=c).value for c in range(29, 42)]
mu = np.array([ws.cell(row=5, column=c).value for c in range(29, 42)], dtype=float)

Sigma = []
for row in range(45, 58):
    fila = [ws.cell(row=row, column=c).value for c in range(31, 44)]
    Sigma.append(fila)
Sigma = np.array(Sigma, dtype=float)

n  = len(tickers)
Rf = 0.0378

bandas = {
    'VOO':  (0.18,  0.26), 'VEA':  (0.09,  0.17), 'VWO':  (0.05,  0.13),
    'VTWO': (0.03,  0.09), 'VCIT': (0.07,  0.14), 'VGIT': (0.06,  0.13),
    'SCHP': (0.04,  0.10), 'HYG':  (0.03,  0.08), 'VNQ':  (0.03,  0.09),
    'IAU':  (0.03,  0.08), 'BIL':  (0.05,  0.05), 'TLT':  (-0.10, -0.05),
    'IGLB': (-0.12, -0.08),
}
w_min = np.array([bandas[t][0] for t in tickers])
w_max = np.array([bandas[t][1] for t in tickers])

# ── Mínima Varianza ───────────────────────────────────────────────────────────
w_mv = cp.Variable(n)
prob_mv = cp.Problem(cp.Minimize(cp.quad_form(w_mv, Sigma)),
                     [cp.sum(w_mv)==1, w_mv>=w_min, w_mv<=w_max])
prob_mv.solve(solver=cp.MOSEK)
std_mv = np.sqrt(prob_mv.value)
ret_mv = w_mv.value @ mu

# ── Retorno máximo factible ───────────────────────────────────────────────────
w_tmp2 = cp.Variable(n)
prob_max = cp.Problem(cp.Maximize(mu @ w_tmp2), [cp.sum(w_tmp2)==1, w_tmp2>=w_min, w_tmp2<=w_max])
prob_max.solve(solver=cp.MOSEK)
ret_max = mu @ w_tmp2.value

print(f"Mínima Varianza: Ret={ret_mv*100:.2f}% Std={std_mv*100:.2f}%")
print(f"Retorno máximo factible: {ret_max*100:.2f}%")

# ── Frontera de 12 puntos (entre min var y max retorno) ───────────────────────
targets = np.linspace(ret_mv, ret_max, 12)
resultados = []
for i, target in enumerate(targets):
    w = cp.Variable(n)
    constraints = [cp.sum(w)==1, w>=w_min, w<=w_max, w@mu==target]
    prob = cp.Problem(cp.Minimize(cp.quad_form(w, Sigma)), constraints)
    prob.solve(solver=cp.MOSEK, verbose=False)
    if prob.status in ["optimal", "optimal_inaccurate"] and prob.value is not None and prob.value > 0:
        std = np.sqrt(prob.value)
        resultados.append({'ret': target, 'std': std, 'w': w.value, 'tipo': 'Frontera'})
        print(f"  Punto {i+1}/12 | Ret: {target*100:.2f}% | Std: {std*100:.2f}%")

# ── Máximo Sharpe (búsqueda fina con más puntos, no solo entre los 12) ───────
targets_finos = np.linspace(ret_mv, ret_max, 200)
mejor_sharpe, mejor_w, mejor_ret, mejor_std = -np.inf, None, None, None
for target in targets_finos:
    w = cp.Variable(n)
    constraints = [cp.sum(w)==1, w>=w_min, w<=w_max, w@mu==target]
    prob = cp.Problem(cp.Minimize(cp.quad_form(w, Sigma)), constraints)
    prob.solve(solver=cp.MOSEK, verbose=False)
    if prob.status in ["optimal", "optimal_inaccurate"] and prob.value is not None and prob.value > 0:
        std = np.sqrt(prob.value)
        sharpe = (target - Rf) / std
        if sharpe > mejor_sharpe:
            mejor_sharpe, mejor_w, mejor_ret, mejor_std = sharpe, w.value, target, std

print(f"\nMáximo Sharpe real: Ret={mejor_ret*100:.2f}% Std={mejor_std*100:.2f}% Sharpe={mejor_sharpe:.4f}")

# ── Agregar Min Var y Max Sharpe como filas extra ─────────────────────────────
resultados.append({'ret': ret_mv, 'std': std_mv, 'w': w_mv.value, 'tipo': 'Mínima Varianza'})
resultados.append({'ret': mejor_ret, 'std': mejor_std, 'w': mejor_w, 'tipo': 'Máximo Sharpe'})

print(f"\nTotal de filas en la tabla: {len(resultados)} (12 frontera + Mín Var + Máx Sharpe)")

# ── Excel ─────────────────────────────────────────────────────────────────────
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Frontera Con Venta en Corto"
headers = ['Tipo', 'Retorno Anualizado', 'Std Anualizada', 'Sharpe Ratio'] + [f'{t} Peso' for t in tickers]
hfill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True)
mv_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
ms_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

for ci, h in enumerate(headers, 1):
    c = ws_out.cell(row=1, column=ci, value=h)
    c.fill, c.font = hfill, hfont
    c.alignment = Alignment(horizontal='center')

for ri, r in enumerate(sorted(resultados, key=lambda x: x['std']), 2):
    ws_out.cell(row=ri, column=1, value=r['tipo'])
    ws_out.cell(row=ri, column=2, value=r['ret']).number_format = '0.00%'
    ws_out.cell(row=ri, column=3, value=r['std']).number_format = '0.00%'
    ws_out.cell(row=ri, column=4, value=f"=(B{ri}-{Rf})/C{ri}").number_format = '0.0000'
    for j in range(n):
        ws_out.cell(row=ri, column=5+j, value=r['w'][j]).number_format = '0.00%'
    if r['tipo'] == 'Mínima Varianza':
        for ci in range(1, 5+n):
            ws_out.cell(row=ri, column=ci).fill = mv_fill
    elif r['tipo'] == 'Máximo Sharpe':
        for ci in range(1, 5+n):
            ws_out.cell(row=ri, column=ci).fill = ms_fill

ws_out.column_dimensions['A'].width = 18
ws_out.column_dimensions['B'].width = 20
ws_out.column_dimensions['C'].width = 18
ws_out.column_dimensions['D'].width = 14
for j in range(n):
    ws_out.column_dimensions[get_column_letter(5+j)].width = 12

wb_out.save('frontera_cv_2.xlsx')
print("\n✅ Excel guardado: frontera_cv_2.xlsx (14 filas: 12 frontera + Mín Var + Máx Sharpe resaltados)")
